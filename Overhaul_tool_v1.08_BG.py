from tkinter import *
from tkinter import Frame,Menu
from tkinter import ttk, filedialog as fd
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
from PIL import ImageTk, Image
from openpyxl.drawing.image import Image as signature
from openpyxl.drawing.image import Image as televic
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from win32com import client
import win32api
import pathlib
import webbrowser as web
from PyPDF2 import PdfFileReader
import time
import os
import win32com.client as win32
import pandas as pd
import shutil
import glob
import random
import configparser


root = Tk() 
root.title("Overhaul tool v1.08")
root.geometry("470x530")
root.iconbitmap("icon.ico")
root.configure(bg='#eff3f3')

#version and changes
def versions():
    root2=Tk()
    root2.title("Версии и промени")
    root2.geometry("400x400")
    root2.iconbitmap("icon.ico")

    label01= Label(root2,text="Version 1.0: Първа Версия")
    label01.pack(pady=10)

    label02= Label(root2,text="Version 1.2: Премахнато полето за LRU \n номер и добавено LRU падащо меню")
    label02.pack(pady=10)

    label03= Label(root2,text="Version 1.4: Добавено падащо меню за избор на име \n адаптирано за ново E15 за 4561 със нов sw.1.06")
    label03.pack(pady=10)

    label04= Label(root2,text="Version 1.05: Добавено репорт файл \n добавено меню бар за версии и IMS база данни")
    label04.pack(pady=10)

    label05= Label(root2,text="Version 1.6: Добавено база данни за 4562 и 4563 (нестабилно !)")
    label05.pack(pady=10)

    label06= Label(root2,text="Version 1.7: Добавена опция за изпращане на имейл\n със документи към клиента")
    label06.pack(pady=10)

    label07= Label(root2,text="Version 1.8: Добавено Админ меню и подобрения в стила")
    label07.pack(pady=10)

    root2.mainloop()
# Database for 33.92.4561 
def database():
    root3=Tk()
    root3.title("IMS база данни")
    root3.geometry("400x320")
    root3.iconbitmap("icon.ico")

    imsentry = ttk.Entry(root3)
    imsentry.pack()

    la1 = Label(root3,text="Note: Ако не намирате информация\n моля обърнете се към инженет или супервайзър", font='Times 9 bold').place(x=60, y=270)

    def search():
            if len(imsentry.get()) == 0:
                la4= Label(root3, text="Празно поле!").place(x=165,y=75)
            else :
                file = glob.glob('Database for 4561/33.92.4561_*'+imsentry.get()+'_*_*_*.txt', recursive=True )

                latest_file = max(file, key=os.path.getctime)

                string_with_brackets = latest_file
                string_without_brackets = string_with_brackets.strip("{}")


                f = open(string_without_brackets)



                la=Label(root3,text=f.readline(100)+ 
                                    f.readline(100)+ 
                                    f.readline(100)+ 
                                    f.readline(100)+ 
                                    f.readline(100)+
                                    f.readline(100)+
                                    f.readline(100)+
                                    f.readline(100)+
                                    f.readline(100)+
                                    f.readline(100)+
                                    f.readline(100)+
                                    f.readline(100)+
                                    f.readline(100))
                                    
                la.place(x=0,y=50)
      
                f.close()



    imsbutton = ttk.Button(root3,text='Търси',command=search)
    imsbutton.pack()

    root3.mainloop()
# Database for 33.92.4562 (unstable!)
def database2():
    root4=Tk()
    root4.title("DSC Box База данни")
    root4.geometry("400x320")
    root4.iconbitmap("icon.ico")

    dsc_box_e = Entry(root4)
    dsc_box_e.pack()

    la1 = Label(root4,text="Note: Ако не намирате информация\n моля обърнете се към инженет или супервайзър", font='Times 9 bold').place(x=60, y=270)

    def search2():
        if len(dsc_box_e.get()) == 0:
            la4= Label(root4, text="Празно поле!").place(x=165,y=75)
        else :
            e1 = ttk.Entry(root4)
            e1.pack_forget()

            file = glob.glob('S:/Productie/U_E15/33.92.4562/archive/33.92.4562_*-'+dsc_box_e.get()+'.xls', recursive=True )
            e1.delete(0,END)
            e1.insert(END,file)
            string_with_brackets = e1.get()
            string_without_brackets = string_with_brackets.strip("{}")


            open_file = pd.read_excel(string_without_brackets,sheet_name='routine test report-1',usecols ='R,W,Z',header=9,nrows=5)


            la=Label(root4,text=file)
            la.place(x=20,y=50)

            la1=Label(root4,text=open_file)
            la1.place(x=70,y=80)

    dsc_box_btn = ttk.Button(root4,text='Търси',command=search2)
    dsc_box_btn.pack()
    root4.mainloop()
# Database for 33.92.4563 (unstable!)
def database3():
    root5=Tk()
    root5.title("SPM Box база данни")
    root5.geometry("400x320")
    root5.iconbitmap("icon.ico")

    spm_box_e = ttk.Entry(root5)
    spm_box_e.pack()

    la1 = Label(root5,text="Note: Ако не намирате информация\n моля обърнете се към инженет или супервайзър", font='Times 9 bold').place(x=60, y=270)

    def search3():
        if len(spm_box_e.get()) == 0:
            la5= Label(root5, text="empty field").place(x=165,y=75)
        else :
            e2 = Entry(root5)
            e2.pack_forget()

            

            file = glob.glob('S:/Productie/U_E15/33.92.4563/archive/33.92.4563_*-'+spm_box_e.get()+'.xls', recursive=True )
            e2.delete(0,END)
            e2.insert(END,file)
            string_with_brackets = e2.get()
            string_without_brackets = string_with_brackets.strip("{}")


            open_file = pd.read_excel(string_without_brackets,sheet_name='routine test report-1',usecols ='R,W,Z',header=9,nrows=5)


            la1=Label(root5,text=file)
            la1.place(x=20,y=50)

            la2=Label(root5,text=open_file)
            la2.place(x=70,y=80)
    
    spm_box_btn = ttk.Button(root5,text='Търси',command=search3)
    spm_box_btn.pack()

    root5.mainloop()
# Send documents
def send_to_client():

    root6=Tk()
    root6.geometry('400x320')
    root6.title('Изпрати към клиента')
    root6.iconbitmap("icon.ico")


    config = configparser.ConfigParser()

    config.read('config.ini')
    set_location=config['User']['location']
    set_email=config['User']['emails']
        

    frame1 = Frame(root6, width=380, height=65, borderwidth=2, relief=GROOVE)
    frame1.place(x=10, y=15)

    frame2 = Frame(root6, width=380, height=65, borderwidth=2, relief=GROOVE)
    frame2.place(x=10, y=85)

    frame3 = Frame(root6, width=380, height=65, borderwidth=2, relief=GROOVE)
    frame3.place(x=10, y=155)

    def send_4561():
        #SEND EMAIL
        os.chdir(set_location+"/Owerhaul Project_4561/EMAIL")
        for file in glob.glob("*.zip"):
            la_info1=Label(root6,text=file+'  Успешно изпратени')
            la_info1.place(x=50,y=250)

        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        #mail.To = 'Sandro.Defilippis@hitachirail.com;domenico.ferrara@hitachirail.com;Lucia.Iadicicco@hitachirail.com;Salvatore.Russo@hitachirail.com;n.arabadjiev@televic.com'
        mail.To = set_email
        mail.Subject = '4809049903' #order number 
        mail.HTMLBody = 'Dear'+'<p>'+' Please find the attached Conformity and routine test report for'+'<p>'+' Purchase order: 4809049903'+'<p>'+' Position : 00030'+'<p>'+' BT material code : 100.360.354' #Email content

        # To attach a file to the email:
        attachments = set_location+'/Owerhaul Project_4561/EMAIL'+'/'+file
        mail.Attachments.Add(attachments)

        mail.Send()

        time.sleep(1)
        os.chdir(set_location+"/Owerhaul Project_4561/EMAIL")
        for file in glob.glob("*.zip"):
            original1 = set_location+'/Owerhaul Project_4561/EMAIL/'+file
            target1 = set_location+'/Owerhaul Project_4561/SENT/'+file 
            shutil.move(original1, target1)       

    def zip_4561():
        os.chdir(set_location+"/Owerhaul Project_4561/EMAIL")
        for file in glob.glob("*.zip"):
            print(file)
        shutil.make_archive('33.92.4561_Documents_'+time.strftime('%d'+'.'+'%m'+'.'+'%Y'), 'zip', set_location+'/Owerhaul Project_4561/TO BE SEND')
        
        os.chdir(set_location+"/Owerhaul Project_4561/BUFFER")
        for file in glob.glob("*.zip"):
            original = set_location+'/Owerhaul Project_4561/BUFFER/'+file
            target = set_location+'/Owerhaul Project_4561/EMAIL/'+file   
            shutil.move(original, target)
            print(file)


        os.chdir(set_location+"/Owerhaul Project_4561/TO BE SEND")
        for file in glob.glob("*.pdf"):
            original_pdf = set_location+'/Owerhaul Project_4561/TO BE SEND/'+file
            target_pdf = set_location+'/Owerhaul Project_4561/PDF/'+file
            shutil.move(original_pdf, target_pdf)
            print(file)

    

    btn1=ttk.Button(root6, text='Изпрати', command=send_4561)
    btn1.place(x=300,y=50)
    la1=ttk.Label(root6, text='Изпрати 33.92.4561 файлове към клиента')
    la1.place(x=60, y=52)

    btn2=ttk.Button(root6,text='Създай', command=zip_4561)
    btn2.place(x=300,y=20)
    la2=ttk.Label(root6, text='Създай 33.92.4561 архив')
    la2.place(x=60, y=22)

    def send_4562():

        os.chdir(set_location+"/Owerhaul Project_4562/EMAIL")
        for file in glob.glob("*.zip"):
            la_info1=Label(root6,text=file+'  Успешно изпратени')
            la_info1.place(x=50,y=250)


        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        #mail.To = 'Sandro.Defilippis@hitachirail.com;domenico.ferrara@hitachirail.com;Lucia.Iadicicco@hitachirail.com;Salvatore.Russo@hitachirail.com;n.arabadjiev@televic.com'
        mail.To = set_email
        mail.Subject = '4809049903'
        mail.HTMLBody = 'Dear'+'<p>'+' Please find the attached Conformity and routine test report for'+'<p>'+' Purchase order: 4809049903'+'<p>'+' Position : 00010'+'<p>'+' BT material code : 100.361.156' #Email content

        # To attach a file to the email (optional):
        attachments = set_location+'/Owerhaul Project_4562/EMAIL'+'/'+file
        mail.Attachments.Add(attachments)

        os.chdir(set_location+"/Owerhaul Project_4562/EMAIL")
        for file in glob.glob("*.zip"):
            original2 = set_location+'/Owerhaul Project_4562/EMAIL/'+file
            target2 = set_location+'/Owerhaul Project_4562/SENT/'+file 
            shutil.move(original2, target2)
        mail.Send()


    def zip_4562():
        os.chdir(set_location+"/Owerhaul Project_4562/EMAIL")
        for file in glob.glob("*.zip"):
            print(file)
        shutil.make_archive('33.92.4562_Documents_'+time.strftime('%d'+'.'+'%m'+'.'+'%Y'), 'zip', set_location+'/Owerhaul Project_4562/TO BE SEND')

        os.chdir(set_location+"/Owerhaul Project_4562/BUFFER")
        for file in glob.glob("*.zip"):
            original = set_location+'/Owerhaul Project_4562/BUFFER/'+file
            target = set_location+'/Owerhaul Project_4562/EMAIL/'+file   
            shutil.move(original, target)
            print(file)


        os.chdir(set_location+"/Owerhaul Project_4562/TO BE SEND")
        for file in glob.glob("*.pdf"):
            original_pdf = set_location+'/Owerhaul Project_4562/TO BE SEND/'+file
            target_pdf = set_location+'/Owerhaul Project_4562/PDF/'+file
            shutil.move(original_pdf, target_pdf)
            print(file)



    btn3=ttk.Button(root6, text='Изпрати', command=send_4562)
    btn3.place(x=300,y=120)
    la3=ttk.Label(root6, text='Изпрати 33.92.4562 файлове към клиента')
    la3.place(x=60, y=122)

    btn4=ttk.Button(root6,text='Създай', command=zip_4562)
    btn4.place(x=300,y=90)
    la4=ttk.Label(root6, text='Създай 33.92.4562 архив')
    la4.place(x=60, y=92)


    def send_4563():

        os.chdir(set_location+"/Owerhaul Project_4563/EMAIL")
        for file in glob.glob("*.zip"):
            la_info3=Label(root6,text=file+'  Успешно изпратени')
            la_info3.place(x=50,y=250)


        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        #mail.To = 'Sandro.Defilippis@hitachirail.com;domenico.ferrara@hitachirail.com;Lucia.Iadicicco@hitachirail.com;Salvatore.Russo@hitachirail.com;n.arabadjiev@televic.com'
        mail.To = set_email
        mail.Subject = '4809049903'
        mail.HTMLBody = 'Dear'+'<p>'+' Please find the attached Conformity and routine test report for'+'<p>'+' Purchase order: 4809049903'+'<p>'+' Position : 00020'+'<p>'+' BT material code : 100.360.114' #Email content

        # To attach a file to the email (optional):
        attachments = set_location+'/Owerhaul Project_4563/EMAIL'+'/'+file
        mail.Attachments.Add(attachments)

        os.chdir(set_location+"/Owerhaul Project_4563/EMAIL")
        for file in glob.glob("*.zip"):
            original3 = set_location+'/Owerhaul Project_4563/EMAIL/'+file
            target3 = set_location+'/Owerhaul Project_4563/SENT/'+file   
            shutil.move(original3, target3)

        mail.Send()


    def zip_4563():
        os.chdir(set_location+"/Owerhaul Project_4563/EMAIL")
        for file in glob.glob("*.zip"):
            print(file)
        shutil.make_archive('33.92.4563_Documents_'+time.strftime('%d'+'.'+'%m'+'.'+'%Y'), 'zip', set_location+'/Owerhaul Project_4563/TO BE SEND')

        os.chdir(set_location+"/Owerhaul Project_4563/BUFFER")
        for file in glob.glob("*.zip"):
            original = set_location+'/Owerhaul Project_4563/BUFFER/'+file
            target = set_location+'/Owerhaul Project_4563/EMAIL/'+file   
            shutil.move(original, target)
            print(file)


        os.chdir(set_location+"/Owerhaul Project_4563/TO BE SEND")
        for file in glob.glob("*.pdf"):
            original_pdf = set_location+'/Owerhaul Project_4563/TO BE SEND/'+file
            target_pdf = set_location+'/Owerhaul Project_4563/PDF/'+file
            shutil.move(original_pdf, target_pdf)
            print(file)


    btn5=ttk.Button(root6, text='Изпрати', command=send_4563)
    btn5.place(x=300,y=190)
    la5=ttk.Label(root6, text='Изпрати 33.92.4563 файлове към клиента')
    la5.place(x=60, y=192)

    btn6=ttk.Button(root6,text='Създай', command=zip_4563)
    btn6.place(x=300,y=160)
    la6=ttk.Label(root6, text='Създай 33.92.4563 архив')
    la6.place(x=60, y=162)
    
    root6.mainloop()
#Admin page
def admin():
    root7=Tk()
    root7.geometry('600x420')
    root7.title('Admin page')
    root7.iconbitmap("icon.ico")

    location_label=ttk.Label(root7,text='Изберете базова локация').place(x=5,y=20)
    location_entry=ttk.Entry(root7,width=50)
    location_entry.place(x=150,y=20)


    email_label=ttk.Label(root7,text='Сменете имейл на клиента').place(x=5,y=60)
    email_entry=ttk.Entry(root7,width=48)
    email_entry.place(x=160,y=60)

    def location():
            location_ask = fd.askdirectory()
            location_entry.delete(0,END)
            location_entry.insert(END, location_ask)

            config = configparser.ConfigParser()

            config.read('config.ini')

            config['User']['location'] = location_entry.get()

            location_set_success_label=Label(root7,text='Path Successfully set').pack()

            with open('config.ini', 'w') as configfile:
                config.write(configfile)

    def email():

        config = configparser.ConfigParser()

        config.read('config.ini')

        config['User']['emails'] = email_entry.get()

        location_set_success_label=Label(root7,text='Email Successfully set').pack()

        with open('config.ini', 'w') as configfile:
            config.write(configfile)


    location_entry_btn=ttk.Button(root7,text='Задай',width=15,command=location).place(x=470,y=18)

    email_entry_btn=ttk.Button(root7,text='Задай',width=15,command=email).place(x=470,y=58)

    root7.mainloop()

menubar = Menu(root)
root.config(menu=menubar)

menubar.add_cascade(
    label="Версии",
    command=versions
)

menubar.add_cascade(
    label="4561 База",
    command=database

)

menubar.add_cascade(
    label="4562 База",
    command=database2

)

menubar.add_cascade(
    label="4563 База",
    command=database3

)

menubar.add_cascade(
    label='Изпрати файлове',
    command = send_to_client)

menubar.add_cascade(
    label='Админ',
    command = admin)


# create a function for askdirectory (for Directory to get original E15's)
def directory():
    a = fd.askdirectory(title="Директория оригинални Е15")
    entry1.delete(0,END)
    entry1.insert(END,a)


# use the function to get the location into Entry (which is hidden)
entry1 = ttk.Entry(root)
entry1.pack_forget()


# create a function for askdirectory (for Directory to upload new E15's)
def directory1():
    b = fd.askdirectory(title="Нова директория Е15 ексел .xlsx")
    entry2.delete(0,END)
    entry2.insert(END,b)

# use the function to get the location into Entry (which is hidden)
entry2 = ttk.Entry(root)
entry2.pack_forget()


# create a function for askdirectory (for Directory to upload new PDF)
def directory2():
    c = fd.askdirectory(title="Нова директория Е15 .pdf")
    entry3.delete(0,END)
    entry3.insert(END, c)


# use the function to get the location into Entry (which is hidden))
entry3 = ttk.Entry(root)
entry3.pack_forget()


def pdf():
    instruction_config = configparser.ConfigParser()

    instruction_config.read('config.ini')
    instruction_location=instruction_config['User']['location']

    web.open(instruction_location+'/Work_instruction.pdf')

# All the action happens here in one function
def myClick():
    wb = load_workbook(entry1.get() + "/" + click.get() + "_" + batch.get() + "-" + serial.get() + ".xlsx")
    ws = wb["routine test report-1"]
    ws1 = wb["conformity"]
    img = signature("Nikolai.png")
    img.width = 170
    img.height = 70

    right_border = Border(right=Side(style='thin'))
    left_border = Border(left=Side(style='thin'))

    # For 33.92.4561
    if click.get() == "33.92.4561":


        wb["routine test report-1"]['S32'] = ""
        wb["routine test report-1"]['S36'] = ""
        wb["routine test report-1"]['S40'] = ""
        wb["routine test report-1"]['S44'] = ""
        #Removes averything from Y and R cell
        wb["routine test report-1"]['Y32'] = ""
        wb["routine test report-1"]['R32'] = ""
        wb["routine test report-1"]['Y36'] = ""
        wb["routine test report-1"]['R36'] = ""
        wb["routine test report-1"]['Y40'] = ""
        wb["routine test report-1"]['R40'] = ""
        wb["routine test report-1"]['Y44'] = ""
        wb["routine test report-1"]['R44'] = ""
        wb["routine test report-1"]['Y48'] = ""
        wb["routine test report-1"]['R48'] = ""
        #DONE!
        wb['routine test report-1']['E14'].font = Font(size=7)
        wb['routine test report-1']['E15'].font = Font(size=7)

        wb["routine test report-1"]['Q32'] = "Done & Approved by:"
        wb["routine test report-1"]['C33'] = "According ITP n° 33.92.4561_ITP phase n° 1, 4 & 5"

        wb["routine test report-1"]['Q36'] = "Done & Approved by:"
        wb["routine test report-1"]['C37'] = "According ITP n° 33.92.4561_ITP phases n°6"

        wb["routine test report-1"]['Q40'] = "Done & Approved by:"
        wb["routine test report-1"]['X40'] = "  N.A."

        wb["routine test report-1"]['Q44'] = "Done & Approved by:"
        wb["routine test report-1"]['C45'] = "According ITP n° 33.92.4561_ITP phases n°6"

        wb["routine test report-1"]['B47'] = "Assembly:"
        wb["routine test report-1"]['C48'] = "Procedure:"
        wb["routine test report-1"]['I48'] = "33.92.4511_E07"
        wb["routine test report-1"]['C49'] = "According ITP n° 33.92.4561_ITP phases n°1-2-3-4"
        wb["routine test report-1"]['Q48'] = "Done & Approved by:"
        wb["routine test report-1"]['X48'] = date.get() + "," + acronym_select.get()

        wb["conformity"]['O12'] = ""
        wb["conformity"]['O13'] = ""
        wb["conformity"]['O14'] = ""
        wb["conformity"]['O16'] = ""
        wb["conformity"]['O18'] = ""
        wb["conformity"]['O22'] = ""
        wb["conformity"]['C20'] = ""
        wb["conformity"]['C21'] = ""
        wb["conformity"]['D39'] = ""
        wb["conformity"]['D40'] = ""
        wb["conformity"]['O39'] = ""
        wb["conformity"]['N39'] = ""
        wb["conformity"]['O41'] = ""
        wb["conformity"]['D43'] = ""
        wb["conformity"]['D44'] = ""
        wb["conformity"]['O43'] = ""
        wb["conformity"]['N43'] = ""
        wb["conformity"]['O45'] = ""
        wb["conformity"]['I51'] = ""
        wb["conformity"]['I52'] = ""

        wb["conformity"]['C20'] = "Control plan reference"
        wb["conformity"]['C21'] = "Riferimento control plan"

        ws1.add_image(img, "P49")

        for ws1 in wb:
            img1 = televic('televic_logo.png')
            ws1.add_image(img1, "B3")

        for ws in wb:
            img1 = televic('televic_logo.png')
            ws.add_image(img1, "B3")

        wb["conformity"]['O12'] = "Hitachi Rail STS stabil. Napoli"
        wb["conformity"]['O13'] = "Via Argine, 425"
        wb["conformity"]['O14'] = "80147 Napoli (NA) - Italia"
        wb["conformity"]['O16'] = "4809049903"
        wb["conformity"]['O18'] = "04.08.2022"
        wb["conformity"]['O20'] = "33_92_4561_ITP"
        wb["conformity"]['O22'] = "33.92.4511_E07"
        wb['conformity']['O45'].alignment = Alignment(horizontal='left')
        wb['conformity']['O45'].alignment = Alignment(vertical='top')
        wb["conformity"]['O41'] = "1"
        wb["conformity"]['O45'] = batch.get() + "-" + serial.get()

        wb["conformity"]['I51'] = "Nikolai Arabadjiev"
        wb["conformity"]['I52'] = conformity_date.get()

        wb.active = wb['routine test report-1']

        wb.save(entry2.get() + "/" + click.get() + "_" + batch.get() + "-" + serial.get() + ".xlsx")

        # export to pdf for 4561
        excel_file = entry2.get() + "/" + click.get() + "_" + batch.get() + "-" + serial.get() + ".xlsx"
        pdf_file = entry3.get() + "/" + click.get() + "_" + batch.get() + "-" + serial.get() + ".pdf"
        excel_path = str(pathlib.Path.cwd() / excel_file)
        pdf_path = str(pathlib.Path.cwd() / pdf_file)

        excel = client.DispatchEx("Excel.Application")
        excel.Visible = 0

        wb = excel.Workbooks.Open(excel_path)
        ws = wb.Worksheets[0]

        wb.SaveAs(pdf_path, FileFormat=57)
        wb.Close()
        excel.Quit()

        upload1 = Label(root, text='33.92.4561_'+batch.get()+'-'+serial.get()+'   Качено!')
        upload1.place(x=155,y=215)
        #Added report file


        report = open('Reports.txt','a')
        report.write('33.92.4561_'+batch.get()+'-'+serial.get()+'   Assembled by:'+acronym_select.get()+'  In Date: '+ date.get()+"    Uploaded:  "+time.strftime('%H'+':'+'%M'+':'+'%S'+'     '+'%d'+'/'+'%m'+'/'+'%Y'+'\n'))
        report.close()


    

    # For 33.92.4562
    if click.get() == "33.92.4562":

        wb["routine test report-1"]['S28'] = ""
        wb["routine test report-1"]['S32'] = ""
        wb["routine test report-1"]['S36'] = ""
        wb["routine test report-1"]['S40'] = ""

        wb["routine test report-1"]['Q28'] = "Done & Approved by:"
        wb["routine test report-1"]['C29'] = "According ITP n° 33.92.4562_ITP phase n° 1, 4 & 5"

        wb["routine test report-1"]['Q32'] = "Done & Approved by:"
        wb["routine test report-1"]['C33'] = "According ITP n° 33.92.4562_ITP phases n°6"

        wb["routine test report-1"]['Q36'] = "Done & Approved by:"
        wb["routine test report-1"]['C37'] = ""
        wb["routine test report-1"]['X36'] = "  N.A."

        wb["routine test report-1"]['Q40'] = "Done & Approved by:"
        wb["routine test report-1"]['C41'] = "According ITP n° 33.92.4562_ITP phases n°6"

        wb["routine test report-1"]['B43'] = "Assembly:"
        wb["routine test report-1"]['C44'] = "Procedure:"
        wb["routine test report-1"]['I44'] = "33.92.4512_E07"
        wb["routine test report-1"]['C45'] = "According ITP n° 33.92.4562_ITP phases n°1-2-3-4"
        wb["routine test report-1"]['Q44'] = "Done & Approved by:"
        wb["routine test report-1"]['X44'] = date.get() + "," + acronym_select.get()

        wb["conformity"]['O12'] = ""
        wb["conformity"]['O13'] = ""
        wb["conformity"]['O14'] = ""
        wb["conformity"]['O16'] = ""
        wb["conformity"]['O18'] = ""
        wb["conformity"]['O22'] = ""
        wb["conformity"]['C20'] = ""
        wb["conformity"]['C21'] = ""

        wb["conformity"]['D39'] = ""
        wb["conformity"]['D40'] = ""

        wb["conformity"]['O39'] = ""
        wb["conformity"]['N39'] = ""

        wb["conformity"]['O41'] = ""

        wb["conformity"]['D43'] = ""
        wb["conformity"]['D44'] = ""

        wb["conformity"]['O43'] = ""
        wb["conformity"]['N43'] = ""

        wb["conformity"]['O45'] = ""
        wb["conformity"]['I53'] = ""
        wb["conformity"]['I54'] = ""

        ws1.add_image(img, "P52")

        for ws1 in wb:
            img1 = televic('televic_logo.png')
            ws1.add_image(img1, "B3")
            
        for ws in wb:
            img1 = televic('televic_logo.png')
            ws.add_image(img1, "B3")

        wb["conformity"]['C20'] = "Control plan reference"
        wb["conformity"]['C21'] = "Riferimento control plan"

        wb["conformity"]['O12'] = "Hitachi Rail STS stabil. Napoli"
        wb["conformity"]['O13'] = "Via Argine, 425"
        wb["conformity"]['O14'] = "80147 Napoli (NA) - Italia"
        wb["conformity"]['O16'] = "4809049903"
        wb["conformity"]['O18'] = "04.08.2022"
        wb["conformity"]['O20'] = "33_92_4562_ITP"
        wb["conformity"]['O22'] = "33.92.4512_E07"
        wb['conformity']['O45'].alignment = Alignment(horizontal='left')
        wb['conformity']['O45'].alignment = Alignment(vertical='top')
        wb["conformity"]['O41'] = "1"
        wb["conformity"]['O45'] = batch.get() + "-" + serial.get()

        wb["conformity"]['I53'] = "Nikolai Arabadjiev"
        wb["conformity"]['I54'] = conformity_date.get()
        wb.active = wb['conformity']

        wb.save(entry2.get() + "/" + click.get() + "_" + batch.get() + "-" + serial.get() + ".xlsx")

        # export to pdf for 4562
        excel_file = entry2.get() + "/" + click.get() + "_" + batch.get() + "-" + serial.get() + ".xlsx"
        pdf_file = entry3.get() + "/" + click.get() + "_" + batch.get() + "-" + serial.get() + ".pdf"
        excel_path = str(pathlib.Path.cwd() / excel_file)
        pdf_path = str(pathlib.Path.cwd() / pdf_file)

        excel = client.DispatchEx("Excel.Application")
        excel.Visible = 0

        wb = excel.Workbooks.Open(excel_path)
        ws = wb.Worksheets[0]

        wb.SaveAs(pdf_path, FileFormat=57)
        wb.Close()
        excel.Quit()

        upload2 = Label(root, text='33.92.4562_'+batch.get()+'-'+serial.get()+'   Качено!')
        upload2.place(x=155,y=215)

        report = open('Reports.txt','a')
        report.write('33.92.4562_'+batch.get()+'-'+serial.get()+'   Assembled by:'+acronym_select.get()+'  In Date: '+ date.get()+"    Uploaded:  "+time.strftime('%H'+':'+'%M'+':'+'%S'+'     '+'%d'+'/'+'%m'+'/'+'%Y'+'\n'))
        report.close()





    # For 33.92.4563
    if click.get() == "33.92.4563":

        wb["routine test report-1"]['S27'] = ""
        wb["routine test report-1"]['S31'] = ""
        wb["routine test report-1"]['S35'] = ""
        wb["routine test report-1"]['S39'] = ""

        wb["routine test report-1"]['Q27'] = "Done & Approved by:"
        wb["routine test report-1"]['C28'] = "According ITP n° 33.92.4563_ITP phase n° 1, 4 & 5"

        wb["routine test report-1"]['Q31'] = "Done & Approved by:"
        wb["routine test report-1"]['C32'] = "According ITP n° 33.92.4563_ITP phases n°6"

        wb["routine test report-1"]['Q35'] = "Done & Approved by:"
        wb["routine test report-1"]['C36'] = ""
        wb["routine test report-1"]['X35'] = "  N.A."

        wb["routine test report-1"]['Q39'] = "Done & Approved by:"
        wb["routine test report-1"]['C40'] = "According ITP n° 33.92.4563_ITP phases n°6"

        wb['routine test report-1'].insert_rows(42,2)
        wb["routine test report-1"]['B42'].border = left_border
        wb["routine test report-1"]['B43'].border = left_border
        wb["routine test report-1"]['AF42'].border = right_border
        wb["routine test report-1"]['AF43'].border = right_border

        ws1.add_image(img, "P52")

        for ws1 in wb:
            img1 = televic('televic_logo.png')
            ws1.add_image(img1, "B3")
            
        for ws in wb:
            img1 = televic('televic_logo.png')
            ws.add_image(img1, "B3")


        wb["routine test report-1"]['B42'] = "Assembly:"
        wb["routine test report-1"]['C43'] = "Procedure:"
        wb["routine test report-1"]['I43'] = "33.92.4513_E07"
        wb["routine test report-1"]['C44'] = "According ITP n° 33.92.4563_ITP phases n°1-2-3-4"
        wb["routine test report-1"]['Q43'] = "Done & Approved by:"
        wb["routine test report-1"]['X43'] = date.get() + "," + acronym_select.get()

        wb["conformity"]['O12'] = ""
        wb["conformity"]['O13'] = ""
        wb["conformity"]['O14'] = ""
        wb["conformity"]['O16'] = ""
        wb["conformity"]['O18'] = ""
        wb["conformity"]['O22'] = ""
        wb["conformity"]['C20'] = ""
        wb["conformity"]['C21'] = ""

        wb["conformity"]['D39'] = ""
        wb["conformity"]['D40'] = ""

        wb["conformity"]['O39'] = ""
        wb["conformity"]['N39'] = ""

        wb["conformity"]['O41'] = ""

        wb["conformity"]['D43'] = ""
        wb["conformity"]['D44'] = ""

        wb["conformity"]['O43'] = ""
        wb["conformity"]['N43'] = ""

        wb["conformity"]['O45'] = ""
        wb["conformity"]['I53'] = ""
        wb["conformity"]['I54'] = ""

        wb["conformity"]['C20'] = "Control plan reference"
        wb["conformity"]['C21'] = "Riferimento control plan"

        wb["conformity"]['O12'] = "Hitachi Rail STS stabil. Napoli"
        wb["conformity"]['O13'] = "Via Argine, 425"
        wb["conformity"]['O14'] = "80147 Napoli (NA) - Italia"
        wb["conformity"]['O16'] = "4809049903"
        wb["conformity"]['O18'] = "04.08.2022"
        wb["conformity"]['O20'] = "33_92_4563_ITP"
        wb["conformity"]['O22'] = "33.92.4513_E07"
        wb['conformity']['O45'].alignment = Alignment(horizontal='left')
        wb['conformity']['O45'].alignment = Alignment(vertical='top')
        wb["conformity"]['O41'] = "1"
        wb["conformity"]['O45'] = batch.get() + "-" + serial.get()

        wb["conformity"]['I53'] = "Nikolai Arabadjiev"
        wb["conformity"]['I54'] = conformity_date.get()
        wb.active = wb['conformity']

        wb.save(entry2.get() + "/" + click.get() + "_" + batch.get() + "-" + serial.get() + ".xlsx")

        # export to pdf for 4563
        excel_file = entry2.get() + "/" + click.get() + "_" + batch.get() + "-" + serial.get() + ".xlsx"
        pdf_file = entry3.get() + "/" + click.get() + "_" + batch.get() + "-" + serial.get() + ".pdf"
        excel_path = str(pathlib.Path.cwd() / excel_file)
        pdf_path = str(pathlib.Path.cwd() / pdf_file)

        excel = client.DispatchEx("Excel.Application")
        excel.Visible = 0

        wb = excel.Workbooks.Open(excel_path)
        var = wb.Worksheets[0]

        wb.SaveAs(pdf_path, FileFormat=57)
        wb.Close()
        excel.Quit()

        upload3 = Label(root, text='33.92.4563_'+batch.get()+'-'+serial.get()+'   Качено!')
        upload3.place(x=155,y=215)


        report = open('Reports.txt','a')
        report.write('33.92.4563_'+batch.get()+'-'+serial.get()+'   Assembled by:'+acronym_select.get()+'  In Date: '+ date.get()+"    Uploaded:  "+time.strftime('%H'+':'+'%M'+':'+'%S'+'     '+'%d'+'/'+'%m'+'/'+'%Y'+'\n'))
        report.close()



# Tempalates
options = [
    "Изберете LRU",
    "33.92.4561",
    "33.92.4562",
    "33.92.4563",
]

# Commands
frame_menu = Frame(root, width=128, height=27, borderwidth=2, relief=GROOVE)
frame_menu.place(x=118, y=17)
click = StringVar(root)
click.set('   Select LRU   ')
drop = ttk.OptionMenu(root,click,*options)
drop.place(x=120, y=19)
Label(root, text="Изберете LRU:").place(x=18, y=18)

#Operators
operators = [
 ' SMO',
 ' MV',
 ' MEM',
 ' TNI',
 ' RZ',
 ' RA', 
 ' RZ',
 ' NIPE',
 ' SP',
 

]

# Dropdown menu for selected people 

acronym_select = ttk.Combobox(root,value=operators,width=9)
acronym_select.current(0)
acronym_select.set('')
acronym_select.place(x=380, y=58)

# LRU Entry if extra work is needed ;) 
#LRU = Entry(root, width=20)
#LRU.pack()


# Batch Entry
batch = ttk.Entry(root, width=20)
batch.place(x=120, y=60)
Label(root, text="Номер поръчка:").place(x=18, y=62)

# Serial Entry
serial = ttk.Entry(root, width=10)
serial.place(x=120, y=100)
Label(root, text="Сериен номер:").place(x=20, y=102)

# Asembly Date
date = ttk.Entry(root, width=12)
date.place(x=380, y=20)
Label(root, text="Дата сглобяване:").place(x=275, y=22)
date.insert(0, "DD/MM/YY")
date.configure(state=DISABLED)


def on_click(event):
    date.configure(state=NORMAL)
    date.delete(0, END)


date.bind("<Button-1>", on_click)

# Asembly Entry changed with dropdown menu only for selected people 
#asem = Entry(root, width=5)
#asem.place(x=380, y=58)
Label(root, text="Сглобено от:").place(x=290, y=60)

# Conformity Date
conformity_date = ttk.Entry(root, width=12)
conformity_date.place(x=380, y=100)
Label(root, text="Дата проверка:").place(x=285, y=102)
conformity_date.insert(0, "DD/MM/YY")
conformity_date.configure(state=DISABLED)


def on_click(event):
    conformity_date.configure(state=NORMAL)
    conformity_date.delete(0, END)


conformity_date.bind("<Button-1>", on_click)

# Upload button
btn = ttk.Button(root, text="Старт", command=myClick)
btn.pack(padx=10, pady=170,ipady=5,ipadx=800)

# Read PDF instruction
readpdf = ttk.Button(root, text="Инструкция", command=pdf)
readpdf.place(x=10,y=470)

# Image for Logo
image1 = Image.open("logo.png")
test = ImageTk.PhotoImage(image1)
label1 = Label(root, image=test, width=200, height=200)
label1.image = test
label1.place(x=130, y=250)

# directory button

btn = ttk.Button(root, text="Оригинална локация", width=20, command=directory).place(x=10, y=135)

# xlsx directory button
btn1 = ttk.Button(root, text="Към локация(xlsx)", width=20, command=directory1).place(x=170, y=135)

# pdf directory button
btn2 = ttk.Button(root, text="Към локация(pdf)", width=20, command=directory2).place(x=330, y=135)


# Dev info

Developer_info = Label(root, text="M.Hasanov Televic GSP .® "+time.strftime('%Y'), font=("courier", 8)).place(x=230, y=480)

root.mainloop()
